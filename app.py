import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import pandas as pd
import logging
import re
from datetime import datetime
from pathlib import Path

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="Me Latte Café — Control de Asistencias WhatsApp", 
    page_icon="☕", 
    layout="wide"
)

# Setup de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Rutas de archivos
DATA_DIR = Path("data")
PHOTOS_DIR = Path("asistencias_fotos")
EXCEL_PATH = DATA_DIR / "registro_asistencias_me_latte_cafe.xlsx"

# Crear directorios si no existen
DATA_DIR.mkdir(exist_ok=True)
PHOTOS_DIR.mkdir(exist_ok=True)

# ==================== FUNCIONES AUXILIARES ====================

def sanitize_filename(filename: str) -> str:
    """Sanitizar nombres de archivo para evitar vulnerabilidades."""
    return re.sub(r'[^a-zA-Z0-9._-]', '_', filename)

def asegurar_archivo_excel():
    """Asegurar que existe el archivo Excel con estructura correcta."""
    if not EXCEL_PATH.exists():
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registro Semanal"
            
            # Estilos
            title_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
            title_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="D2691E", end_color="D2691E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Título
            title_row = ws.append(["ME LATTE CAFÉ — CONTROL DE ASISTENCIAS"])
            ws.merge_cells('A1:J1')
            ws['A1'].fill = title_fill
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Subtítulo
            ws.append(["Semana en curso"])
            ws.merge_cells('A2:J2')
            ws['A2'].font = Font(italic=True)
            
            # Filas vacías
            ws.append([])
            ws.append([])
            ws.append([])
            
            # Encabezados
            headers = ["Día", "Fecha", "Nombre", "Puesto", "H. Entrada", "F. Entrada", "H. Salida", "F. Salida", "Retardo", "Observaciones"]
            header_row = ws.append(headers)
            
            # Aplicar estilos a encabezados
            for cell in ws[6]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 30
            
            wb.save(EXCEL_PATH)
            logger.info(f"Archivo Excel creado en {EXCEL_PATH}")
            return True
        except Exception as e:
            logger.error(f"Error al crear archivo Excel: {e}")
            st.error(f"❌ Error al crear el archivo: {e}")
            return False
    return True

def employee_checked_in_today(nombre: str, date_str: str) -> bool:
    """Verificar si el empleado ya registró asistencia hoy."""
    try:
        if not EXCEL_PATH.exists():
            return False
            
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Registro Semanal"]
        
        for r in range(7, ws.max_row + 1):
            name_cell = ws.cell(row=r, column=3).value
            date_cell = ws.cell(row=r, column=2).value
            
            if name_cell and name_cell.strip() == nombre.strip() and date_cell == date_str:
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        logger.error(f"Error verificando entrada duplicada: {e}")
        return False

def guardar_asistencia(selected_emp: str, uploaded_photo, empleados_db: dict) -> tuple[bool, str]:
    """
    Guardar asistencia en Excel y foto.
    Retorna: (éxito: bool, mensaje: str)
    """
    try:
        if not uploaded_photo:
            return False, "Por favor, sube una fotografía."
        
        if selected_emp not in empleados_db:
            return False, "Empleado no válido."
        
        now = datetime.now()
        date_str = now.strftime("%d/%m/%Y")
        time_str = now.strftime("%H:%M:%S")
        dia_nombre = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"][now.weekday()]
        
        emp_info = empleados_db[selected_emp]
        nombre = selected_emp.split(" (")[0]
        puesto = emp_info["puesto"]
        
        # Verificar entrada duplicada
        if employee_checked_in_today(nombre, date_str):
            return False, f"⚠️ {nombre} ya registró asistencia hoy a las {date_str}."
        
        # Guardar foto con nombre sanitizado
        filename_base = sanitize_filename(f"{emp_info['tel']}_{now.strftime('%Y%m%d_%H%M%S')}")
        filename = PHOTOS_DIR / f"{filename_base}.jpg"
        
        try:
            with open(filename, "wb") as f:
                f.write(uploaded_photo.getbuffer())
            logger.info(f"Foto guardada: {filename}")
        except Exception as e:
            logger.error(f"Error guardando foto: {e}")
            return False, f"Error al guardar la fotografía: {e}"
        
        # Guardar en Excel
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Registro Semanal"]
            
            # Encontrar primera fila vacía (sin datos en columna "Nombre")
            target_row = None
            for r in range(7, 1000):  # Aumentado límite
                val = ws.cell(row=r, column=3).value
                if val is None or str(val).strip() == "":
                    target_row = r
                    break
            
            if not target_row:
                wb.close()
                return False, "No se encontró espacio disponible en el registro."
            
            # Asegurar que la fila existe
            while ws.max_row < target_row:
                ws.append([])
            
            # Llenar datos
            ws.cell(row=target_row, column=1, value=dia_nombre)
            ws.cell(row=target_row, column=2, value=date_str)
            ws.cell(row=target_row, column=3, value=nombre)
            ws.cell(row=target_row, column=4, value=puesto)
            ws.cell(row=target_row, column=5, value=time_str)
            ws.cell(row=target_row, column=6, value="✔ Verificado (WhatsApp)")
            ws.cell(row=target_row, column=10, value=f"Evidencia: {filename.name}")
            
            # Aplicar alineación
            for col in range(1, 11):
                ws.cell(row=target_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            
            wb.save(EXCEL_PATH)
            wb.close()
            logger.info(f"Asistencia registrada para {nombre} en fila {target_row}")
            
            return True, f"✅ ¡Asistencia de {nombre} registrada a las {time_str}!"
        
        except Exception as e:
            logger.error(f"Error guardando en Excel: {e}")
            return False, f"Error al guardar en Excel: {e}"
    
    except Exception as e:
        logger.error(f"Error en guardar_asistencia: {e}")
        return False, f"Error inesperado: {e}"

def cargar_registros() -> pd.DataFrame:
    """Cargar registros desde Excel."""
    try:
        if not EXCEL_PATH.exists():
            return pd.DataFrame()
        
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Registro Semanal"]
        
        data = []
        for r in range(7, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
            # Solo incluir filas que tienen al menos nombre
            if row_vals[2]:  # Columna "Nombre"
                data.append(row_vals)
        
        wb.close()
        
        if not data:
            return pd.DataFrame()
        
        df = pd.DataFrame(data, columns=["Día", "Fecha", "Nombre", "Puesto", "H. Entrada", "F. Entrada", "H. Salida", "F. Salida", "Retardo", "Observaciones"])
        return df
    
    except Exception as e:
        logger.error(f"Error cargando registros: {e}")
        st.error(f"❌ Error al cargar registros: {e}")
        return pd.DataFrame()

# ==================== INICIALIZACIÓN ====================

asegurar_archivo_excel()

# ==================== INTERFAZ ====================

st.title("☕ Me Latte Café — Panel de Control de Asistencias por WhatsApp")
st.markdown("Esta aplicación simula y gestiona la recepción de fotografías de asistencia enviadas por los colaboradores.")

# Base de datos de empleados
empleados_db = {
    "Carlos Ruiz (Barista)": {"tel": "5219931234567", "puesto": "Barista"},
    "Ana Gómez (Cajera)": {"tel": "5219939876543", "puesto": "Cajera"},
    "Luis Pérez (Cocinero)": {"tel": "5219935554433", "puesto": "Cocinero (Chilaquiles)"},
    "María López (Mesera)": {"tel": "5219932221100", "puesto": "Mesera"}
}

# Sidebar - Simulador
st.sidebar.header("📱 Simulador de Entrada (WhatsApp)")

if empleados_db:
    selected_emp = st.sidebar.selectbox("Seleccionar Empleado:", list(empleados_db.keys()))
    uploaded_photo = st.sidebar.file_uploader("Sube la fotografía (Selfie de asistencia)", type=["jpg", "jpeg", "png"])
    
    col1, col2 = st.sidebar.columns(2)
    
    with col1:
        if st.button("✅ Registrar Asistencia", type="primary", use_container_width=True):
            success, mensaje = guardar_asistencia(selected_emp, uploaded_photo, empleados_db)
            
            if success:
                st.sidebar.success(mensaje)
                st.rerun()
            else:
                st.sidebar.warning(mensaje)
    
    with col2:
        if st.button("🔄 Actualizar", use_container_width=True):
            st.rerun()
else:
    st.sidebar.error("❌ No hay empleados en la base de datos.")

# ==================== VISTA PRINCIPAL ====================

st.subheader("📊 Estado Actual del Registro de Asistencias")

df = cargar_registros()

if not df.empty:
    # Mostrar estadísticas
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("👥 Total de Registros", len(df))
    
    with col2:
        empleados_unicos = df["Nombre"].nunique()
        st.metric("👤 Empleados Únicos", empleados_unicos)
    
    with col3:
        hoy = datetime.now().strftime("%d/%m/%Y")
        registros_hoy = len(df[df["Fecha"] == hoy])
        st.metric("📅 Hoy", registros_hoy)
    
    with col4:
        st.metric("✅ Verificados", len(df[df["F. Entrada"].notna()]))
    
    st.divider()
    
    # Tabla interactiva
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    # Descargar Excel
    try:
        with open(EXCEL_PATH, "rb") as file:
            st.download_button(
                label="📥 Descargar Registro (Excel)",
                data=file,
                file_name=f"asistencias_me_latte_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        logger.error(f"Error en descarga: {e}")
        st.error(f"Error al descargar archivo: {e}")
else:
    st.info("📋 Aún no hay registros en la semana.")

# ==================== FOOTER ====================

st.divider()
col1, col2, col3 = st.columns(3)

with col1:
    st.caption("📍 Me Latte Café")
    
with col2:
    st.caption(f"🕐 Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

with col3:
    st.caption("✨ v2.0 - Mejorado")
